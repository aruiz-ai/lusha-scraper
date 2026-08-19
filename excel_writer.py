import os
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import config

HEADERS = ["Nombre", "Cargo", "Correo", "Teléfono", "URL de LinkedIn"]
COLUMN_WIDTHS = [35, 45, 40, 25, 70]


def sanitize_filename(value):
    value = re.sub(r'[\\/:*?"<>|]', "_", value).strip()
    return value or "empresa"


def export_to_excel(rows, company):
    wb = Workbook()
    ws = wb.active
    ws.title = "Contactos"

    ws.append(HEADERS)
    header_fill = PatternFill("solid", fgColor="0B66C2")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append([
            row.get("name", ""),
            row.get("role", ""),
            "",
            "",
            row.get("url", ""),
        ])

    for rindex in range(2, ws.max_row + 1):
        link = ws.cell(row=rindex, column=5)
        if link.value:
            link.hyperlink = link.value
            link.font = Font(color="0563C1", underline="single")

    for index, width in enumerate(COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width

    ws.freeze_panes = "A2"

    filename = f"{sanitize_filename(company)}.xlsx"
    filepath = os.path.join(config.EXPORTS_DIR, filename)
    wb.save(filepath)
    return filepath, filename